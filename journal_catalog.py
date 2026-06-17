from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from thefuzz import fuzz

from schema import CandidatePaper

DEFAULT_CATALOG_PATH = Path(__file__).resolve().parent / "references" / "custom_journal_catalog.xlsx"
DEFAULT_SHEET_NAME = "分级列表"
FUZZY_THRESHOLD = 95


@dataclass
class JournalCatalogEntry:
    journal_name: str
    issn: str | None
    level_raw: str
    level_group: str


@dataclass
class CatalogMatch:
    entry: JournalCatalogEntry
    match_type: str
    score: int


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    value = unicodedata.normalize("NFKC", str(value)).lower().strip()
    value = re.sub(r"&", "and", value)
    value = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_issn(value: str | None) -> str:
    if not value:
        return ""
    value = str(value).strip().upper()
    return re.sub(r"[^0-9X]", "", value)


def map_level_group(level_raw: str | None) -> str:
    text = (level_raw or "").upper()
    if "A" in text:
        return "A"
    if "B" in text:
        return "B"
    if "C" in text:
        return "C"
    return "Other"


def load_catalog(
    workbook_path: Path = DEFAULT_CATALOG_PATH,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> list[JournalCatalogEntry]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    entries: list[JournalCatalogEntry] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        journal_name = str(row[1]).strip() if row[1] else ""
        issn = normalize_issn(row[2]) if row[2] else None
        level_raw = str(row[3]).strip() if row[3] else "Other"
        if not journal_name:
            continue
        entries.append(
            JournalCatalogEntry(
                journal_name=journal_name,
                issn=issn,
                level_raw=level_raw,
                level_group=map_level_group(level_raw),
            )
        )
    return entries


def match_journal(
    journal: str | None,
    issn: str | None,
    catalog: list[JournalCatalogEntry],
    fuzzy_threshold: int = FUZZY_THRESHOLD,
) -> CatalogMatch | None:
    normalized_issn = normalize_issn(issn)
    normalized_journal = normalize_text(journal)

    if normalized_issn:
        for entry in catalog:
            if entry.issn and entry.issn == normalized_issn:
                return CatalogMatch(entry=entry, match_type="issn_exact", score=100)

    if normalized_journal:
        for entry in catalog:
            if normalize_text(entry.journal_name) == normalized_journal:
                return CatalogMatch(entry=entry, match_type="journal_exact", score=100)

    best_match: CatalogMatch | None = None
    for entry in catalog:
        score = fuzz.token_sort_ratio(normalize_text(entry.journal_name), normalized_journal)
        if score >= fuzzy_threshold and (best_match is None or score > best_match.score):
            best_match = CatalogMatch(entry=entry, match_type="journal_fuzzy", score=score)
    return best_match


def filter_and_label(
    candidates: list[CandidatePaper],
    catalog: list[JournalCatalogEntry],
    fuzzy_threshold: int = FUZZY_THRESHOLD,
) -> tuple[list[tuple[CandidatePaper, CatalogMatch]], list[CandidatePaper]]:
    matched: list[tuple[CandidatePaper, CatalogMatch]] = []
    rejected: list[CandidatePaper] = []
    for candidate in candidates:
        match = match_journal(candidate.journal, candidate.issn, catalog, fuzzy_threshold=fuzzy_threshold)
        if match is None:
            rejected.append(candidate)
            continue
        matched.append((candidate, match))
    return matched, rejected
